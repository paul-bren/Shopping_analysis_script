
import pymysql.cursors
import openpyxl
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
from xlsxwriter.workbook import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time

list1 = []
list2 = []
data = []
xTickMarks = []


#Find file ending in csv. Don't need the following code but FYI incase I ever need it for different projects.
'''
directory = os.path.join('C:/Users/leden/OneDrive/Desktop/Paul Stuff/Scripts/Analysis_proj')
for roots, dirs, files in os.walk(directory):
    for file in files:
       if file.endswith(".csv"):
           f=open(file, 'r')
           test = file
           #  perform calculation
           f.close()'''

wb = openpyxl.load_workbook('C:/Users/leden/OneDrive/Desktop/Paul Stuff/Scripts/Analysis_proj/test2.xlsx')
#ws = wb['sheet']

#This selects the worksheet to do the analysis from. Use a print(ws) to understand
ws = wb.worksheets[-1]

#connect to the database
db = pymysql.connect(host='localhost',    # your host, usually localhost
                     user='',         # your username
                     passwd='',  # your password
                     port ='',
                     database='')        # name of the data base

try:
    #you must create a Cursor object to query a database
    cur = db.cursor()
    cur2 = db.cursor()
except:
    print("Unable to connect to DB. Please check DB is available and try again")
try:
    query1 = """DELETE FROM Items WHERE InsertDate <= NOW()"""
    cur.execute(query1)
except:
    print("Unable to delete data from database")

#NB The following list is important. Please add the shops you visit regularly so the script can run some analysis    
mystring = ['']


for i in range(2, ws.max_row + 1):
    #row = [paul.value for paul in ws[i]] #This gets the value of each cell while looping using 'i' as the index
    insertdate = datetime.now()
    date = ws.cell(i,1).value
    full_transaction = ws.cell(i,2).value
    vendor = ws.cell(i,2).value.upper()
    paul = vendor.split()
    for word in paul:
        if word in mystring:
            vendor = word
            break
        else:
            vendor = "OTHER"
    debit = ws.cell(i,3).value
    if debit is None:
        debit = 0
    credit = ws.cell(i,4).value
    balance = ws.cell(i,5).value
    values = (insertdate, date, full_transaction, vendor, debit, credit, balance)
    try:
        query2 = """INSERT INTO Items (InsertDate, Date, Full_Transaction, Vendor, Debit, Credit, Balance) VALUES (%s, %s, %s, %s, %s, %s, %s)"""
        cur.execute (query2, values)
    except:
        cur.rollback()
    #finally:
        #print ("Unable to insert data. Please try again")

db.commit()
months = ['01','02','03','04','05','06','07','08','09','10','11','12']
days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
while True:
    question1 = input("From what date would you like to analyze from?  ")
    if len(question1) < 10:
        continue
    elif len(question1) == 10:
        pass
    else:
        print("That is not a date. More than 10 chars ")
        continue
    splits = question1.split('-')
    if splits[0] != '2021':
        print("Please look at year ")
        continue
    elif splits[1] not in months:
        print("Please check the month ")
        continue
    elif splits[2] not in days:
        print("Please check the day ")
        continue
    break

query3 = (f"select sum(debit) AS Total, vendor from Items where (date(Date) between  DATE_FORMAT(NOW() ,(%s)) AND NOW() ) group by vendor;")
#explained here -> https://stackoverflow.com/questions/11808232/how-do-i-select-between-the-1st-day-of-the-current-month-and-current-day-in-mysq
#'%Y-%m-01'
try:
    cur.execute(query3, question1)
except:
    print ("Please check the date and try again")
    

query4 = '''select * from Items where vendor = 'OTHER' AND (date(Date) between  DATE_FORMAT(NOW() ,(%s)) AND NOW() ) order by debit DESC;'''
cur2.execute(query4, question1)

list1.append("Please find a break down of your overall spending since " + str(question1) + '''<br>''')
list1.append(""+ '''<br>''')
print("")
print("Please find a break down of your overall spending since " + question1)
print("")
for row in cur.fetchall():
    data.append(int(row[0]))
    xTickMarks.append(str(row[1]))
    main_content = "You have spent " + str(row[0]) + " euro in " + str(row[1]) + " since " + str(question1)
    list1.append(main_content + '''<br>''')
    print("You have spent " + str(row[0]) + " euro in " + str(row[1]) + " since " + str(question1))
list1.append("" + '''<br>''')
list1.append("Please find a break down on category 'OTHER' below" + '''<br>''')
list1.append("" + '''<br>''')
print("")
print("Please find a break down on category 'OTHER' below" )
print("")
for p in cur2.fetchall():
    other_content = "You have spent " + str(p[4]) + " euro in " + str(p[2]) + " since " + str(question1)
    list2.append(other_content + '''<br>''')
    print("You have spent " + str(p[4]) + " euro in " + str(p[2]) + " since " + str(question1))

list2 = list1 + list2

cur.close()
cur2.close()
db.close()

fig = plt.figure()
ax = fig.add_subplot(111)

## necessary variables
ind = np.arange(len(data))                # the x locations for the groups
#print(ind)
width = 0.35                      # the width of the bars

## the bars
rects1 = ax.bar(ind, data, width, color='black', error_kw=dict(elinewidth=2,ecolor='red'))

# axes and labels
ax.set_xlim(-width,len(ind)+width)
#ax.set_xlim(-0.5,len(ind)+width)
try:
    max_val = max(data)
except:
    print("Sorry, looks like there is no data to analyse. Please check the DB for data and try again. ")
    time.sleep(2)
    exit()
#print(max_val)
ax.set_ylim(0, max_val + 200)


ax.set_ylabel('Money spent')
ax.set_xlabel('Shops')
ax.set_title('Shopping Analysis')

#ax.set_xticks(ind+width)
ax.set_xticks(ind)
xtickNames = ax.set_xticklabels(xTickMarks)
plt.setp(xtickNames, rotation=15, fontsize=10)


#See following URL for below: https://stackoverflow.com/questions/53066633/python-how-to-show-values-on-top-of-bar-plot
for i, v in enumerate(data):
    ax.text(i - .25, v + 10, str(v), color='blue', fontweight='bold')


#To get smtplib email working with Gmail you have to enable less secure apps in Gmail
'''
#try the below with mail_box in message.attach(MIMEText(mail_content, 'plain'))

#The mail addresses and password
sender_address = 'paulbrennan99@gmail.com'
sender_pass = ''
receiver_address = 'paulbrennan99@gmail.com'
#Setup the MIME
message = MIMEMultipart() 
message['From'] = sender_address
message['To'] = receiver_address
message['Subject'] = 'Spending analysis'   #The subject line
#The body and the attachments for the mail
message.attach(MIMEText(str(list2),'html'))
#message2.attach(MIMEText(str(list2),'html'))

#message.attach(part1)
#Create SMTP session for sending the mail
session = smtplib.SMTP('smtp.gmail.com', 587) #use gmail with port
session.starttls() #enable security
session.login(sender_address, sender_pass) #login with mail_id and password
text = message.as_string()
session.sendmail(sender_address, receiver_address, text)
print('Mail Sent')
session.quit()'''

plt.show()    



